#! python3

import openpyxl, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

os.chdir(r'H:\python\files\Chapter 13\project_3')

wb = openpyxl.load_workbook('Project_3.xlsx')
ws = wb.active

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

for col in range(1, ws.max_column+1):
    for row in range(1, ws.max_row+1):
        ws_new[get_column_letter(row)+str(col)] = \
                ws[get_column_letter(col)+str(row)].value

bold12 = Font(size=12, bold=True)

for i in range(1, ws.max_column+1):
    ws_new['A'+str(i)].font = bold12

wb_new.save('Project_3_inverted.xlsx')
