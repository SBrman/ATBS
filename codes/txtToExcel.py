#! python3

import openpyxl, os, pprint, re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

os.chdir(r'H:\python\files\Chapter 13\project_4')

textFiles = {} # key=filename, value=[line_1, line_2...]

wb = openpyxl.Workbook()
ws = wb.active

os.chdir('.\\TextFiles')
for file in os.listdir():
    textFiles.setdefault(os.path.basename(file), [])
    text = open(os.path.abspath(file), 'r')
    text_lines = text.readlines()
    textFiles[os.path.basename(file)] = text_lines

fileNum = re.compile(r'(.*_)(\d+)(.txt)')

os.chdir('..\\')

mo = fileNum.search(list(textFiles.keys())[0])
before = mo.group(1)
after = mo.group(3)

    
def keyMaker(num):
    
    return before+str(num)+after


for column in range(1, len(textFiles)+1):
    key_made = keyMaker(column)
    for row in range(2, len(textFiles[key_made])+2):
        ws[get_column_letter(column)+str(row)]= textFiles[key_made][row-2]

for column in range(1, len(textFiles)+1):
    key_made = keyMaker(column)
    ws[get_column_letter(column)+'1'] = key_made
    ws[get_column_letter(column)+'1'].font = Font(size = '14', bold=True)

    if column < int((len(textFiles)) / 2):
        ws.column_dimensions[get_column_letter(column)].width = 60
    else:
        ws.column_dimensions[get_column_letter(column)].width = 120

wb.save('project_4.xlsx')
