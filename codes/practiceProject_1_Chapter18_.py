#! python3

'''Random Chore Assignment Emailer
Write a program that takes a list of people’s email addresses and a list
of chores that need to be done and randomly assigns chores to people.
Email each person their assigned chores. If you’re feeling ambitious, keep
a record of each person’s previously assigned chores so that you can make
sure the program avoids assigning anyone the same chore they did last
time. For another possible feature, schedule the program to run once a
week automatically.'''

import os, sys, copy, random, pprint, csv
from pprint import pformat

os.chdir(r'H:\python\files\Chapter 18\PracticeProject1')
sys.path.append(r'H:\python\files\Chapter 18\PracticeProject1')


def writeFile(dictionary):
    #For saving the people_chores.py file
    file = open('people_chores.py', 'w')
    file.write('people = '+ pformat(dictionary))
    file.close()
    return 1


# importing people_chores file or creating one if that does not exist
if 'people_chores.py' in os.listdir('.'):
    from people_chores import people
else:
    import openpyxl
    wb = openpyxl.load_workbook(r'H:/python/files/Chapter 18/duesRecords.xlsx')
    ws = wb.active

    everything = {}
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row,column=1).value
        email = ws.cell(row=row,column=2).value
        everything[name] = {'email': email, 'chores': []}

    writeFile(everything)
    from people_chores import people


# selecting new chores for each of the people
chores = ['dishes', 'bathroom', 'vacuum', 'walk dog']

for name in people.keys():
    tempChores = copy.copy(chores)
    for chore in people[name]['chores']:
        tempChores.remove(chore)
    people[name]['chores'].append(random.choice(tempChores))
    
writeFile(people)




# for saving details to csv File
csvFile = open('people_chores.csv', 'w', newline='')
writer = csv.writer(csvFile)
writer.writerow(['Name', 'Email']+['Chore %s' %(x+1) for x in range(len(people[name]['chores']))])
for name in people.keys():
    row = [name, people[name]['email']]+[chore for chore in people[name]['chores']]
    writer.writerow(row)
csvFile.close()



# for sending the emails
for name in people.keys():
    print('Recipient: %s\nSubject: Your task is %s for today\nBody:\nHi %s,\n\
I am mailing to let you know about your task for today. Your task is %s for today.\
\nBest,\nMe\n\n'
                    %(people[name]['email'], people[name]['chores'][-1], name, \
                      people[name]['chores'][-1]))


